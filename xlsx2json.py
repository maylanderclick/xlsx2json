import openpyxl
import json
import argparse
import os

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('source', type=str, help='Source xlsx file.')
    parser.add_argument('result', type=str, help='Result name file.')
    parser.add_argument('-a', '--append', action='store_true', help='Append in result file.')
    args = parser.parse_args()
    sc_path = 'test.xlsx'
    res_path = 'test.json'

    if args.source:
        sc_path = args.source
        if not os.path.exists(sc_path):
            raise FileNotFoundError('Scource')

    if args.result:
        res_path = args.result

    if args.append and not os.path.exists(res_path):
        raise FileNotFoundError('Result')

    wb = openpyxl.load_workbook(sc_path)
    ws = wb.active
    dc = dict()
    for i in range(1, len(ws['A'])+1):
        dc.setdefault(ws.cell(row=i, column=1).value, int(ws.cell(row=i, column=2).value))

    if args.append:
        with open(res_path, 'r') as f:
            obj = json.load(f)
            obj.update(dc)
            dc = obj

    with open(res_path, 'w') as f:
        json.dump(dc, f, indent=4)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print('Error: ' + str(e))

        